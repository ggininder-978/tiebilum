import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def generate_excel():
    wb = Workbook()
    
    # 1. Main Sheet
    ws1 = wb.active
    ws1.title = "2025銷售總表"
    
    with open('knowledge/wiki/analysis/2025_sales_sku_master.csv', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            ws1.append(row)
            
    # Styling Header
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 2. Profit Analysis Sheet
    ws2 = wb.create_sheet("產品獲利診斷")
    ws2.append(["產品名稱", "銷售數量", "總營收", "貢獻利潤", "淨利率 (%)"])
    
    # Aggregate data manually
    product_stats = {}
    with open('knowledge/wiki/analysis/2025_sales_sku_master.csv', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = row.get('product_name_std', 'Unknown')
            if not name: continue
            qty = float(row.get('quantity_2025', 0) or 0)
            rev = float(row.get('revenue', 0) or 0)
            profit = float(row.get('contribution_profit', 0) or 0)
            
            if name not in product_stats:
                product_stats[name] = {'qty': 0, 'rev': 0, 'profit': 0}
            product_stats[name]['qty'] += qty
            product_stats[name]['rev'] += rev
            product_stats[name]['profit'] += profit

    # Fill Sheet 2
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    for name, stats in sorted(product_stats.items(), key=lambda x: x[1]['profit'], reverse=True):
        margin = (stats['profit'] / stats['rev'] * 100) if stats['rev'] != 0 else 0
        row_data = [name, stats['qty'], stats['rev'], stats['profit'], round(margin, 2)]
        ws2.append(row_data)
        if margin < 5: # Flag low or negative margin
            ws2.last_row = ws2.max_row
            ws2.cell(row=ws2.max_row, column=5).fill = red_fill

    # Save
    import os
    os.makedirs('outputs', exist_ok=True)
    wb.save('outputs/Tiebilum_2025_Sales_Master.xlsx')
    print("Excel file saved to outputs/Tiebilum_2025_Sales_Master.xlsx")

if __name__ == "__main__":
    generate_excel()
